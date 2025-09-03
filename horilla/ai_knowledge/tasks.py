from celery import shared_task
from django.core.files.storage import default_storage
from django.utils import timezone
from django.conf import settings
import os
import logging
import json
import re
from typing import Dict, List, Any

# Import for document processing
try:
    import PyPDF2
    from docx import Document as DocxDocument
    import openpyxl
    from PIL import Image
    import pytesseract
except ImportError:
    # Handle missing dependencies gracefully
    PyPDF2 = None
    DocxDocument = None
    openpyxl = None
    Image = None
    pytesseract = None

from .models import (
    AIDocument, KnowledgeBaseEntry, TrainingData, 
    AIIntent, DocumentProcessingLog
)

logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Document processing utilities"""
    
    @staticmethod
    def extract_text_from_pdf(file_path: str) -> str:
        """Extract text from PDF file"""
        if not PyPDF2:
            raise ImportError("PyPDF2 is required for PDF processing")
        
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            logger.error(f"Error extracting text from PDF {file_path}: {str(e)}")
            raise
        
        return text.strip()
    
    @staticmethod
    def extract_text_from_docx(file_path: str) -> str:
        """Extract text from DOCX file"""
        if not DocxDocument:
            raise ImportError("python-docx is required for DOCX processing")
        
        try:
            doc = DocxDocument(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text.strip()
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def extract_text_from_xlsx(file_path: str) -> str:
        """Extract text from XLSX file"""
        if not openpyxl:
            raise ImportError("openpyxl is required for XLSX processing")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}\n")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)
                
                text_parts.append("\n")
            
            return "\n".join(text_parts).strip()
        except Exception as e:
            logger.error(f"Error extracting text from XLSX {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def extract_text_from_txt(file_path: str) -> str:
        """Extract text from TXT file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read().strip()
            except Exception as e:
                logger.error(f"Error extracting text from TXT {file_path}: {str(e)}")
                raise
        except Exception as e:
            logger.error(f"Error extracting text from TXT {file_path}: {str(e)}")
            raise

class KnowledgeExtractor:
    """Extract knowledge from processed text"""
    
    @staticmethod
    def extract_qa_pairs(text: str) -> List[Dict[str, str]]:
        """Extract Q&A pairs from text"""
        qa_pairs = []
        
        # Simple pattern matching for Q&A
        qa_pattern = r'(?:Q|Question|\?)\s*:?\s*(.+?)(?:A|Answer|\n)\s*:?\s*(.+?)(?=(?:Q|Question|\?)|$)'
        matches = re.findall(qa_pattern, text, re.IGNORECASE | re.DOTALL)
        
        for question, answer in matches:
            qa_pairs.append({
                'question': question.strip(),
                'answer': answer.strip()
            })
        
        return qa_pairs
    
    @staticmethod
    def extract_procedures(text: str) -> List[Dict[str, Any]]:
        """Extract procedures and step-by-step instructions"""
        procedures = []
        
        # Pattern for numbered steps
        step_pattern = r'(\d+\.\s*.+?)(?=\d+\.|$)'
        steps = re.findall(step_pattern, text, re.DOTALL)
        
        if steps:
            procedures.append({
                'type': 'procedure',
                'steps': [step.strip() for step in steps]
            })
        
        return procedures
    
    @staticmethod
    def extract_policies(text: str) -> List[Dict[str, str]]:
        """Extract policy information"""
        policies = []
        
        # Pattern for policy statements
        policy_keywords = ['policy', 'rule', 'regulation', 'guideline', 'standard']
        
        for keyword in policy_keywords:
            pattern = f'({keyword}[^.]*\.(?:[^.]*\.)*?)'
            matches = re.findall(pattern, text, re.IGNORECASE)
            
            for match in matches:
                policies.append({
                    'type': keyword,
                    'content': match.strip()
                })
        
        return policies

class IntentGenerator:
    """Generate AI intents from processed documents"""
    
    @staticmethod
    def generate_intents_from_qa(qa_pairs: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        """Generate intents from Q&A pairs"""
        intents = []
        
        for i, qa in enumerate(qa_pairs):
            intent_name = f"qa_intent_{i+1}"
            
            # Generate variations of the question
            examples = [qa['question']]
            
            # Simple variations (this could be enhanced with NLP)
            base_question = qa['question'].lower()
            if base_question.startswith('what'):
                examples.append(base_question.replace('what', 'tell me about'))
            elif base_question.startswith('how'):
                examples.append(base_question.replace('how', 'explain how'))
            
            intents.append({
                'name': intent_name,
                'description': f"Intent for: {qa['question'][:50]}...",
                'examples': examples,
                'responses': [qa['answer']],
                'confidence_score': 0.8
            })
        
        return intents

@shared_task(bind=True)
def process_document_async(self, document_id: int):
    """Process uploaded document asynchronously"""
    try:
        document = AIDocument.objects.get(id=document_id)
        
        # Log start of processing
        DocumentProcessingLog.objects.create(
            document=document,
            level='info',
            message='Starting document processing',
            processing_step='initialization'
        )
        
        # Update document status
        document.status = 'processing'
        document.save()
        
        # Get file path
        file_path = document.file.path
        
        # Extract text based on file type
        processor = DocumentProcessor()
        extracted_text = ""
        
        try:
            if document.file_type.lower() == 'pdf':
                extracted_text = processor.extract_text_from_pdf(file_path)
            elif document.file_type.lower() in ['docx', 'doc']:
                extracted_text = processor.extract_text_from_docx(file_path)
            elif document.file_type.lower() in ['xlsx', 'xls']:
                extracted_text = processor.extract_text_from_xlsx(file_path)
            elif document.file_type.lower() == 'txt':
                extracted_text = processor.extract_text_from_txt(file_path)
            else:
                raise ValueError(f"Unsupported file type: {document.file_type}")
            
            # Save extracted text
            document.extracted_text = extracted_text
            document.save()
            
            DocumentProcessingLog.objects.create(
                document=document,
                level='success',
                message=f'Text extraction completed. Extracted {len(extracted_text)} characters.',
                processing_step='text_extraction'
            )
            
        except Exception as e:
            DocumentProcessingLog.objects.create(
                document=document,
                level='error',
                message=f'Text extraction failed: {str(e)}',
                processing_step='text_extraction',
                details={'error': str(e)}
            )
            document.status = 'error'
            document.processing_notes = f'Text extraction failed: {str(e)}'
            document.save()
            return
        
        # Extract knowledge
        try:
            extractor = KnowledgeExtractor()
            
            # Extract Q&A pairs
            qa_pairs = extractor.extract_qa_pairs(extracted_text)
            
            # Create knowledge base entries
            for qa in qa_pairs:
                KnowledgeBaseEntry.objects.create(
                    title=qa['question'][:200],
                    content=qa['answer'],
                    entry_type='faq',
                    source_document=document,
                    confidence_score=0.7
                )
            
            # Extract procedures
            procedures = extractor.extract_procedures(extracted_text)
            for proc in procedures:
                KnowledgeBaseEntry.objects.create(
                    title=f"Procedure from {document.title}",
                    content=json.dumps(proc),
                    entry_type='procedure',
                    source_document=document,
                    confidence_score=0.6
                )
            
            # Extract policies
            policies = extractor.extract_policies(extracted_text)
            for policy in policies:
                KnowledgeBaseEntry.objects.create(
                    title=f"{policy['type'].title()} from {document.title}",
                    content=policy['content'],
                    entry_type='policy',
                    source_document=document,
                    confidence_score=0.8
                )
            
            DocumentProcessingLog.objects.create(
                document=document,
                level='success',
                message=f'Knowledge extraction completed. Created {len(qa_pairs)} Q&A pairs, {len(procedures)} procedures, {len(policies)} policies.',
                processing_step='knowledge_extraction'
            )
            
        except Exception as e:
            DocumentProcessingLog.objects.create(
                document=document,
                level='warning',
                message=f'Knowledge extraction partially failed: {str(e)}',
                processing_step='knowledge_extraction',
                details={'error': str(e)}
            )
        
        # Generate AI intents
        try:
            intent_generator = IntentGenerator()
            qa_pairs = extractor.extract_qa_pairs(extracted_text)
            intents = intent_generator.generate_intents_from_qa(qa_pairs)
            
            for intent_data in intents:
                # Check if intent already exists
                if not AIIntent.objects.filter(name=intent_data['name']).exists():
                    intent = AIIntent.objects.create(
                        name=intent_data['name'],
                        description=intent_data['description'],
                        examples=intent_data['examples'],
                        responses=intent_data['responses'],
                        confidence_score=intent_data['confidence_score']
                    )
                    intent.source_documents.add(document)
            
            DocumentProcessingLog.objects.create(
                document=document,
                level='success',
                message=f'Intent generation completed. Created {len(intents)} intents.',
                processing_step='intent_generation'
            )
            
        except Exception as e:
            DocumentProcessingLog.objects.create(
                document=document,
                level='warning',
                message=f'Intent generation failed: {str(e)}',
                processing_step='intent_generation',
                details={'error': str(e)}
            )
        
        # Generate training data
        try:
            qa_pairs = extractor.extract_qa_pairs(extracted_text)
            
            for i, qa in enumerate(qa_pairs):
                TrainingData.objects.create(
                    name=f"Training data from {document.title} - {i+1}",
                    training_type='intent',
                    input_text=qa['question'],
                    expected_output=qa['answer'],
                    intent_label=f"qa_intent_{i+1}",
                    source_document=document,
                    confidence_threshold=0.7
                )
            
            DocumentProcessingLog.objects.create(
                document=document,
                level='success',
                message=f'Training data generation completed. Created {len(qa_pairs)} training entries.',
                processing_step='training_data_generation'
            )
            
        except Exception as e:
            DocumentProcessingLog.objects.create(
                document=document,
                level='warning',
                message=f'Training data generation failed: {str(e)}',
                processing_step='training_data_generation',
                details={'error': str(e)}
            )
        
        # Mark document as processed
        document.status = 'processed'
        document.processing_notes = 'Document processing completed successfully'
        document.save()
        
        DocumentProcessingLog.objects.create(
            document=document,
            level='success',
            message='Document processing completed successfully',
            processing_step='completion'
        )
        
        return {
            'status': 'success',
            'document_id': document_id,
            'message': 'Document processed successfully'
        }
        
    except AIDocument.DoesNotExist:
        logger.error(f"Document with ID {document_id} not found")
        return {
            'status': 'error',
            'document_id': document_id,
            'message': 'Document not found'
        }
    
    except Exception as e:
        logger.error(f"Error processing document {document_id}: {str(e)}")
        
        try:
            document = AIDocument.objects.get(id=document_id)
            document.status = 'error'
            document.processing_notes = f'Processing failed: {str(e)}'
            document.save()
            
            DocumentProcessingLog.objects.create(
                document=document,
                level='error',
                message=f'Document processing failed: {str(e)}',
                processing_step='error',
                details={'error': str(e)}
            )
        except:
            pass
        
        return {
            'status': 'error',
            'document_id': document_id,
            'message': str(e)
        }

@shared_task
def cleanup_old_logs(days: int = 30):
    """Clean up old processing logs"""
    from datetime import timedelta
    
    cutoff_date = timezone.now() - timedelta(days=days)
    deleted_count = DocumentProcessingLog.objects.filter(
        created_at__lt=cutoff_date
    ).delete()[0]
    
    logger.info(f"Cleaned up {deleted_count} old processing logs")
    return deleted_count

@shared_task
def update_knowledge_base_scores():
    """Update confidence scores for knowledge base entries"""
    entries = KnowledgeBaseEntry.objects.all()
    updated_count = 0
    
    for entry in entries:
        # Simple scoring based on content length and keywords
        content_score = min(len(entry.content) / 1000, 1.0)
        keyword_score = len(entry.keywords.split(',')) * 0.1 if entry.keywords else 0
        
        new_score = min((content_score + keyword_score) / 2, 1.0)
        
        if abs(entry.confidence_score - new_score) > 0.1:
            entry.confidence_score = new_score
            entry.save()
            updated_count += 1
    
    logger.info(f"Updated confidence scores for {updated_count} knowledge base entries")
    return updated_count

@shared_task
def generate_training_data_from_feedback(feedback_data: Dict[str, Any]):
    """Generate training data from user feedback"""
    try:
        # Process feedback and create training data
        # This is a placeholder for actual feedback processing logic
        logger.info(f"Processing feedback: {feedback_data}")
        
        # Implementation would depend on feedback structure
        # For now, just log the feedback
        
    except Exception as e:
        logger.error(f"Error processing feedback: {str(e)}")
        raise
    
    return {'status': 'success', 'message': 'Feedback processed'}


@shared_task(bind=True)
def start_training_process(self, training_data_id: int):
    """Start the training process for a specific training data item"""
    try:
        from .models import TrainingData
        
        # Get the training data
        training_data = TrainingData.objects.get(id=training_data_id)
        
        logger.info(f"Starting training process for: {training_data.name}")
        
        # Update status to processing
        training_data.training_progress = 10
        training_data.training_stage = 'preprocessing'
        training_data.save()
        
        # Simulate preprocessing phase
        import time
        time.sleep(2)
        
        # Update to data validation phase
        training_data.training_progress = 25
        training_data.training_stage = 'validating'
        training_data.save()
        
        # Validate training data
        if not training_data.input_text or not training_data.expected_output:
            raise ValueError("Training data is incomplete")
        
        time.sleep(2)
        
        # Update to model training phase
        training_data.training_progress = 50
        training_data.training_stage = 'training'
        training_data.save()
        
        # Simulate model training
        time.sleep(3)
        
        # Update to evaluation phase
        training_data.training_progress = 75
        training_data.training_stage = 'evaluating'
        training_data.save()
        
        # Simulate evaluation
        time.sleep(2)
        
        # Complete training
        training_data.training_progress = 100
        training_data.training_stage = 'completed'
        training_data.training_completed_at = timezone.now()
        training_data.save()
        
        logger.info(f"Training completed for: {training_data.name}")
        
        return {
            'status': 'success',
            'message': f'Training completed for {training_data.name}',
            'training_data_id': training_data_id,
            'final_progress': 100
        }
        
    except TrainingData.DoesNotExist:
        logger.error(f"Training data with ID {training_data_id} not found")
        return {
            'status': 'error',
            'message': f'Training data with ID {training_data_id} not found'
        }
    except Exception as e:
        logger.error(f"Error in training process for ID {training_data_id}: {str(e)}")
        
        # Update training data with error status
        try:
            training_data = TrainingData.objects.get(id=training_data_id)
            training_data.training_stage = 'failed'
            training_data.save()
        except:
            pass
        
        return {
            'status': 'error',
            'message': f'Training failed: {str(e)}',
            'training_data_id': training_data_id
        }


@shared_task
def batch_training_process(training_data_ids: List[int]):
    """Process multiple training data items in batch"""
    results = []
    
    for training_id in training_data_ids:
        try:
            result = start_training_process.delay(training_id)
            results.append({
                'training_data_id': training_id,
                'task_id': result.id,
                'status': 'started'
            })
        except Exception as e:
            logger.error(f"Failed to start training for ID {training_id}: {str(e)}")
            results.append({
                'training_data_id': training_id,
                'status': 'failed',
                'error': str(e)
            })
    
    return {
        'status': 'success',
        'message': f'Batch training started for {len(training_data_ids)} items',
        'results': results
    }